"""Map Scopus author positions to OpenAlex author IDs via DOI matching.

For each researcher's Scopus publications:
  1. Query Scopus for the paper's full author list
  2. Find the researcher's position in that list
  3. Look up the corresponding OpenAlex author by position
"""

import argparse
import ast
import logging
import os
import shutil
import time
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook

from src.config import settings

logger = logging.getLogger(__name__)

INPUT_DIR = "publicaciones_scopus_excel"
OUTPUT_DIR = "publicaciones_openalex_per_researcher"
OPENALEX_PREFIX = "https://api.openalex.org/works/"


def scopus_lookup_by_doi(doi: str, api_key: str) -> dict | None:
    """Query Scopus for a paper's author list by DOI."""
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "Accept": "application/json"}
    params = {
        "query": f'DOI("{doi.strip()}")',
        "count": 1,
        "view": "COMPLETE",
        "field": "dc:title,prism:doi,author",
    }
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        logger.warning("Scopus error %d for DOI %s", response.status_code, doi)
        return None
    entries = response.json().get("search-results", {}).get("entry", [])
    return entries[0] if entries else None


def openalex_for_doi_by_index(
    doi: str, seq: int, pause: float = 0.5
) -> tuple[str | None, str | None]:
    """Look up OpenAlex work by DOI and return the author at a given position.

    Args:
        doi: Paper DOI
        seq: 1-based author position from Scopus
        pause: Rate-limiting pause between requests

    Returns:
        (openalex_work_id, openalex_author_id) tuple
    """
    doi_url = "https://doi.org/" + doi.strip()
    lookup_url = OPENALEX_PREFIX + requests.utils.quote(doi_url, safe=":/")
    resp = requests.get(lookup_url)
    if resp.status_code != 200:
        logger.warning("OpenAlex error %d for DOI %s", resp.status_code, doi)
        return None, None

    data = resp.json()
    work_id = data.get("id")
    authorships = data.get("authorships", [])
    idx = seq - 1
    author_id = None
    if 0 <= idx < len(authorships):
        author_id = authorships[idx].get("author", {}).get("id")

    time.sleep(pause)
    return work_id, author_id


def process_researcher_files(
    input_dir: str = INPUT_DIR,
    output_dir: str = OUTPUT_DIR,
    clear_dir: bool = False,
) -> None:
    """Process all researcher Scopus files and map to OpenAlex IDs.

    For each researcher Excel file, matches Scopus author positions
    to OpenAlex author IDs via shared DOIs.
    """
    api_key = settings.scopus_api_key

    if clear_dir and os.path.isdir(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    for fname in os.listdir(input_dir):
        if not fname.lower().endswith((".xlsx", ".xls")):
            continue

        base, _ = os.path.splitext(fname)
        parts = base.split("_")
        scopus_id = parts[-1]
        researcher_label = "_".join(parts[:-1])
        out_filename = f"{researcher_label}_{scopus_id}.xlsx"
        out_path = os.path.join(output_dir, out_filename)

        if os.path.exists(out_path):
            logger.info("[Skip] %s exists", out_filename)
            continue

        df_in = pd.read_excel(os.path.join(input_dir, fname))
        rows = []

        for _, row in df_in.iterrows():
            doi = row.get("doi") or row.get("prism:doi")
            title = row.get("dc:title") or row.get("title")
            if not doi or pd.isna(doi):
                continue

            scopus_entry = scopus_lookup_by_doi(doi, api_key)
            if not scopus_entry:
                continue
            raw_auth = scopus_entry.get("author")
            if not raw_auth:
                continue
            authors = ast.literal_eval(raw_auth) if isinstance(raw_auth, str) else raw_auth

            seq = None
            for a in authors:
                if a.get("authid") == scopus_id:
                    seq = int(a.get("@seq", a.get("seq")))
                    break
            if seq is None:
                continue

            work_id, oa_author = openalex_for_doi_by_index(doi, seq)
            rows.append({
                "title": title,
                "doi": doi,
                "scopus_position": seq,
                "openalex_work_id": work_id,
                "openalex_author_id": oa_author,
            })

        df_out = pd.DataFrame(rows)
        if df_out.empty:
            logger.info("No valid papers for %s", researcher_label)
            continue

        df_out["unique_author_first_appearance"] = ~df_out["openalex_author_id"].duplicated()

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, sheet_name="Sheet1", index=False, startrow=2)

        wb = load_workbook(out_path)
        ws = wb.active
        ws["A1"] = f"Researcher: {researcher_label}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df_out.shape[1])
        wb.save(out_path)

        logger.info("Wrote %d rows for %s -> %s", len(df_out), researcher_label, out_filename)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--clear-dir", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO)
    process_researcher_files(clear_dir=args.clear_dir)
