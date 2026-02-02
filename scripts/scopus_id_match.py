# -*- coding: utf-8 -*-
# Made by : Juan Ceresa
"""
Overview:
  1. Loop through each researcher file in the input folder.
  2. For each researcher, read their list of DOIs from the Scopus-generated spreadsheet.
  3. For each DOI:
       a. Query Scopus API to get the paper’s full author list.
       b. Find this researcher’s position in that author list.
       c. Query OpenAlex by DOI and pick the corresponding author entry by position.
  4. Build a consolidated table per researcher with columns:
       - Paper title, DOI,
       - Author position in Scopus (scopus_position),
       - OpenAlex work ID and OpenAlex author ID,
       - Flag `unique_author_first_appearance` marking the first time this author appears in the list.
  5. Save one Excel per researcher, with the researcher’s name at the top and the table starting on row 3.

Usage:
  python scopus_openalex_mapping.py [--clear-dir]

Options:
  --clear-dir    If provided, clear and recreate the output folder before running.
"""

import os
import shutil
import time
import ast
import argparse
import requests
import pandas as pd
from openpyxl import load_workbook

# ─── CONFIGURATION ─────────────────────────────────────────────────────────┐
INPUT_DIR       = "publicaciones_scopus_excel"            # where Scopus spreadsheets live
OUTPUT_DIR      = "publicaciones_openalex_per_researcher"
SCOPUS_API_KEY  = ""  # insert your Scopus API key here
OPENALEX_PREFIX = "https://api.openalex.org/works/"          # base URL for DOI lookups
# ────────────────────────────────────────────────────────────────────────────┘


def scopus_lookup_by_doi(doi: str, api_key: str) -> dict | None:
    """
    Given a DOI, call Scopus with view=COMPLETE to get the paper’s authors.
    Returns the first matching record or None on failure.
    """
    url    = "https://api.elsevier.com/content/search/scopus"
    hdrs   = {"X-ELS-APIKey": api_key, "Accept": "application/json"}
    params = {
        "query": f'DOI("{doi.strip()}")',
        "count": 1,
        "view":  "COMPLETE",
        "field": "dc:title,prism:doi,author"
    }
    response = requests.get(url, headers=hdrs, params=params)
    if response.status_code != 200:
        print(f"[Scopus] Error {response.status_code} looking up DOI {doi}")
        return None
    results = response.json()
    entries = results.get("search-results", {}).get("entry", [])
    return entries[0] if entries else None


def openalex_for_doi_by_index(doi: str, seq: int, pause: float = 0.5) -> tuple[str|None, str|None]:
    """
    Given a DOI and an author sequence number, call OpenAlex to:
      1. Retrieve the work by DOI.
      2. Use the sequence index to pick the matching author’s ID in the OpenAlex authorships.
    Returns (openalex_work_id, openalex_author_id).
    """
    # Build the OpenAlex URL by first converting DOI -> https://doi.org/<doi>
    doi_url = "https://doi.org/" + doi.strip()
    lookup_url = OPENALEX_PREFIX + requests.utils.quote(doi_url, safe=':/')
    resp = requests.get(lookup_url)
    if resp.status_code != 200:
        print(f"[OpenAlex] Error {resp.status_code} looking up DOI {doi}")
        return None, None

    data = resp.json()
    work_id = data.get("id")
    authorships = data.get("authorships", [])

    # OpenAlex uses 0-based indexing internally, Scopus gives 1-based @seq
    idx = seq - 1
    author_id = None
    if 0 <= idx < len(authorships):
        author_id = authorships[idx].get("author", {}).get("id")

    time.sleep(pause)  # rate-limit friendly pause
    return work_id, author_id


if __name__ == "__main__":
    # Parse --clear-dir option
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--clear-dir",
        action="store_true",
        help="Delete and recreate OUTPUT_DIR before running"
    )
    args = parser.parse_args()

    # Prepare output directory (clear if requested)
    if args.clear_dir:
        if os.path.isdir(OUTPUT_DIR):
            shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Process each researcher file
    for fname in os.listdir(INPUT_DIR):
        if not fname.lower().endswith(('.xlsx', '.xls')):
            continue

        # Derive researcher label and scopus ID from filename
        base, _ = os.path.splitext(fname)
        parts = base.split("_")
        scopus_id = parts[-1]
        researcher_label = "_".join(parts[:-1])
        out_filename = f"{researcher_label}_{scopus_id}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, out_filename)

        # Skip files already processed
        if os.path.exists(out_path):
            print(f"[Skip] {out_filename} exists, skipping")
            continue

        # Read the input table of DOIs and titles
        df_in = pd.read_excel(os.path.join(INPUT_DIR, fname))
        rows = []  # accumulate result rows here

        # Loop through each paper entry
        for _, row in df_in.iterrows():
            doi = row.get("doi") or row.get("prism:doi")
            title = row.get("dc:title") or row.get("title")
            if not doi or pd.isna(doi):
                continue  # no DOI, skip

            # Fetch Scopus list of authors for this DOI
            scopus_entry = scopus_lookup_by_doi(doi, SCOPUS_API_KEY)
            if not scopus_entry:
                continue
            raw_auth = scopus_entry.get("author")
            if not raw_auth:
                continue

            # Convert the author list into Python list of dicts
            authors = ast.literal_eval(raw_auth) if isinstance(raw_auth, str) else raw_auth

            # Find this researcher’s position in that list
            seq = None
            for a in authors:
                if a.get("authid") == scopus_id:
                    seq = int(a.get("@seq", a.get("seq")))
                    break
            if seq is None:
                continue  # researcher not on this paper

            # Lookup corresponding author ID in OpenAlex
            work_id, oa_author = openalex_for_doi_by_index(doi, seq)
            rows.append({
                "title": title,
                "doi": doi,
                "scopus_position": seq,
                "openalex_work_id": work_id,
                "openalex_author_id": oa_author
            })

        # Build DataFrame of results
        df_out = pd.DataFrame(rows)
        if df_out.empty:
            print(f"[!] No valid papers for {researcher_label}, skipping")
            continue

        # Mark the first appearance of each author within this list
        df_out["unique_author_first_appearance"] = ~df_out["openalex_author_id"].duplicated()

        # Write the results to a new Excel file
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # Data table starts on row 3 (0-indexed startrow=2)
            df_out.to_excel(writer, sheet_name="Sheet1", index=False, startrow=2)

        # Add a header in the saved file: "Researcher: <label>" in cell A1
        wb = load_workbook(out_path)
        ws = wb.active
        ws["A1"] = f"Researcher: {researcher_label}"
        # Merge across all columns
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df_out.shape[1])
        wb.save(out_path)

        print(f"[✓] Wrote {len(df_out)} rows for {researcher_label} → {out_filename}")
